"""
General Ledger parser — converts a detailed GL (transaction-level) into
TB-equivalent aggregated balances so `analyze_trial_balance` can process it
without knowing the difference.

Target format today is Zoho Books "Detailed General Ledger" export. Common
header variants from Tally / QuickBooks / Busy are also recognised.

Also preserves the raw transactions so the QoE page can mine customer /
vendor concentration without re-parsing the file.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook


# Header synonym maps. Keys are lower-cased, stripped of non-alphanumerics
# so "Account Name" matches "account_name" matches "AccountName".
_ACCOUNT_NAME_KEYS = {
    "accountname", "account", "ledger", "ledgername", "particulars",
    "glaccount", "accountdescription",
}
_DEBIT_KEYS = {"debit", "debitamount", "dr", "netdebit", "debitvalue"}
_CREDIT_KEYS = {"credit", "creditamount", "cr", "netcredit", "creditvalue"}
_CONTACT_KEYS = {
    "transactiondetails", "contactname", "party", "partyname",
    "customer", "vendor", "supplier", "description", "narration",
}
_DATE_KEYS = {"date", "transactiondate", "voucherdate", "txndate"}
_TXN_TYPE_KEYS = {"transactiontype", "voucher", "vouchertype", "type"}
_TXN_ID_KEYS = {"transactionid", "voucherno", "voucher", "txnid", "reference"}
_NET_KEYS = {"netamount", "net", "balance"}


def _normalize(key: str) -> str:
    """'Account  Name' -> 'accountname'; drops spaces/underscores/dashes."""
    return "".join(ch for ch in (key or "").lower() if ch.isalnum())


def _parse_number(value: Any) -> float:
    """Tolerate commas, currency symbols, parentheses (accountant negatives),
    and stray whitespace. Return 0.0 for empty / unparseable inputs."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return 0.0
    # Parentheses as negative
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    s = s.replace(",", "").replace("₹", "").replace("Rs.", "").replace("Rs", "")
    s = s.replace("$", "").replace(" ", "")
    try:
        n = float(s)
    except ValueError:
        return 0.0
    return -n if neg else n


@dataclass
class GLRow:
    """One transaction line from the GL. We keep the raw party name so QoE
    can compute customer/vendor concentration downstream."""
    account_name: str
    debit: float
    credit: float
    contact: str = ""
    date: str = ""
    txn_type: str = ""
    txn_id: str = ""
    group_hint: str = ""  # "Income" / "Expense" / ... — Zoho emits these as section banners


@dataclass
class GLParseResult:
    """What the GL parser returns to the dispatcher.

    tb_entries feeds straight into `analyze_trial_balance`; transactions
    stays around so QoE can mine it for customer concentration, top
    vendors, cash flow patterns, etc. (wired in Wave 6)."""
    tb_entries: list[dict] = field(default_factory=list)
    transactions: list[dict] = field(default_factory=list)
    total_rows: int = 0
    source_format: str = "unknown"
    warnings: list[str] = field(default_factory=list)


def parse_gl_excel(content: bytes) -> GLParseResult:
    """Parse a Zoho Books / Tally / QuickBooks General Ledger Excel export.

    Strategy:
      1. Load workbook, pick the first non-empty worksheet (usually
         "Detailed General Ledger" for Zoho).
      2. Scan first 20 rows for a header row — look for the presence of an
         account_name-like column AND a debit-or-credit column.
      3. Iterate rows, tracking section banners ("Income", "Expense", etc.)
         as group hints.
      4. For each transaction, accumulate per-account net debit / credit
         balances (this is how we flatten a GL into a TB).
    """
    result = GLParseResult()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        result.warnings.append(f"Could not open workbook: {e}")
        return result

    # Pick the sheet most likely to contain the GL. Most Zoho exports only
    # have one sheet, but Tally can stash multiple — we pick the one with
    # the most rows that look like transactions.
    best_sheet = None
    best_score = -1
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Cheap scoring: sheets with "general ledger", "gl", or "ledger" in
        # the name win. Fall back to row count.
        name_l = sheet_name.lower()
        score = 0
        if "general ledger" in name_l or "gl" in name_l or "ledger" in name_l:
            score += 1000
        score += ws.max_row or 0
        if score > best_score:
            best_score = score
            best_sheet = ws
    if best_sheet is None:
        result.warnings.append("Workbook has no usable sheets.")
        return result

    ws = best_sheet

    # Scan up to 20 header rows. Many Zoho exports have 4-6 banner rows
    # (company name, report title, date range, etc.) before the header.
    header_row = -1
    col_map: dict[str, int] = {}
    for row_idx in range(1, min(21, (ws.max_row or 0) + 1)):
        row = [(c.value if c is not None else "") for c in ws[row_idx]]
        norm = [_normalize(str(c)) for c in row]
        account_col = next(
            (i for i, v in enumerate(norm) if v in _ACCOUNT_NAME_KEYS), -1
        )
        debit_col = next((i for i, v in enumerate(norm) if v in _DEBIT_KEYS), -1)
        credit_col = next((i for i, v in enumerate(norm) if v in _CREDIT_KEYS), -1)
        if account_col == -1 or (debit_col == -1 and credit_col == -1):
            continue
        header_row = row_idx
        col_map["account"] = account_col
        if debit_col != -1:
            col_map["debit"] = debit_col
        if credit_col != -1:
            col_map["credit"] = credit_col
        col_map["contact"] = next(
            (i for i, v in enumerate(norm) if v in _CONTACT_KEYS), -1
        )
        col_map["date"] = next(
            (i for i, v in enumerate(norm) if v in _DATE_KEYS), -1
        )
        col_map["txn_type"] = next(
            (i for i, v in enumerate(norm) if v in _TXN_TYPE_KEYS), -1
        )
        col_map["txn_id"] = next(
            (i for i, v in enumerate(norm) if v in _TXN_ID_KEYS), -1
        )
        col_map["net"] = next(
            (i for i, v in enumerate(norm) if v in _NET_KEYS), -1
        )
        result.source_format = _detect_format(norm)
        break

    if header_row == -1:
        result.warnings.append(
            "Could not locate a GL header row. Expected columns like "
            "'Account Name' + 'Debit' / 'Credit'."
        )
        return result

    # Build aggregated TB-equivalent balances keyed by account name.
    per_account_debit: dict[str, float] = {}
    per_account_credit: dict[str, float] = {}
    per_account_group: dict[str, str] = {}
    transactions: list[dict] = []

    current_group = ""
    section_names = {
        "income", "revenue", "sales", "other income",
        "expense", "expenses", "direct expenses", "indirect expenses",
        "assets", "current assets", "non-current assets", "fixed assets",
        "liabilities", "current liabilities", "non-current liabilities",
        "equity", "equities", "capital accounts",
    }

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        cells = list(row)
        if not cells:
            continue

        raw_name = ""
        if col_map["account"] < len(cells):
            raw_name = str(cells[col_map["account"]] or "").strip()
        if not raw_name:
            continue

        # Section banner rows are named like "Income" or "Current Assets".
        name_l = raw_name.lower()
        if name_l in section_names:
            current_group = raw_name
            continue
        # Skip Zoho's subtotal rows ("Total for Sales", "Total").
        if name_l.startswith("total for") or name_l == "total":
            continue
        # Skip the opening balance marker if present.
        if name_l in {"opening balance", "ob"}:
            continue

        def _cell(key: str, default: Any = "") -> Any:
            idx = col_map.get(key, -1)
            if idx < 0 or idx >= len(cells):
                return default
            return cells[idx] if cells[idx] is not None else default

        debit = _parse_number(_cell("debit", 0))
        credit = _parse_number(_cell("credit", 0))
        # Some exports only carry a signed "net amount" column.
        if debit == 0 and credit == 0:
            net = _parse_number(_cell("net", 0))
            if net > 0:
                debit = net
            elif net < 0:
                credit = -net
        if debit == 0 and credit == 0:
            continue

        contact = str(_cell("contact", "") or "").strip()
        date = str(_cell("date", "") or "").strip()
        txn_type = str(_cell("txn_type", "") or "").strip()
        txn_id = str(_cell("txn_id", "") or "").strip()

        per_account_debit[raw_name] = per_account_debit.get(raw_name, 0.0) + debit
        per_account_credit[raw_name] = per_account_credit.get(raw_name, 0.0) + credit
        if current_group and raw_name not in per_account_group:
            per_account_group[raw_name] = current_group

        transactions.append({
            "account_name": raw_name,
            "debit": debit,
            "credit": credit,
            "contact": contact,
            "date": date,
            "txn_type": txn_type,
            "txn_id": txn_id,
            "group": current_group,
        })

    # Collapse per-account totals into net debit / credit (a GL can have
    # both sides populated for the same account across its transactions; a
    # TB shows the net closing balance on whichever side it falls).
    for account_name in per_account_debit.keys() | per_account_credit.keys():
        total_debit = per_account_debit.get(account_name, 0.0)
        total_credit = per_account_credit.get(account_name, 0.0)
        net = total_debit - total_credit
        entry = {
            "account_name": account_name,
            "debit": net if net > 0 else 0.0,
            "credit": -net if net < 0 else 0.0,
            "group": per_account_group.get(account_name, ""),
        }
        result.tb_entries.append(entry)

    result.transactions = transactions
    result.total_rows = len(transactions)

    if not result.tb_entries:
        result.warnings.append(
            "No transactions with non-zero debit or credit were found."
        )
    return result


def _detect_format(normalized_headers: list[str]) -> str:
    """Best-effort detection of which accounting system produced the file."""
    header_set = set(normalized_headers)
    # Zoho Books GL has this very specific combo
    if "transactiondetails" in header_set and "netamount" in header_set:
        return "zoho_books"
    if "vouchertype" in header_set or "voucher" in header_set:
        return "tally"
    if "transactiontype" in header_set and "reference" in header_set:
        return "quickbooks"
    return "generic_gl"


def customer_concentration(transactions: list[dict], top_n: int = 10) -> list[dict]:
    """Aggregate sales-side transactions by contact and return the top N
    customers by revenue. Used by QoE.

    Two detection paths (tried in order):

    1. **Sales / Revenue accounts.** If any transaction is booked to an
       account whose name contains "sales", "revenue", or "income",
       aggregate the credit side of those transactions by contact. This
       is the clean canonical path (what Tally exports typically look
       like).

    2. **Accounts Receivable fallback.** Some Zoho Books setups post
       invoices directly into Accounts Receivable with the revenue side
       flowing to Finished Goods or Inventory. In that case, AR debits
       (invoice creation) are a good proxy for customer revenue and we
       group those by contact instead.
    """
    def _bucket(filter_fn) -> tuple[dict[str, float], float]:
        buckets: dict[str, float] = {}
        total = 0.0
        for txn in transactions:
            if not filter_fn(txn):
                continue
            contact = (txn.get("contact") or "").strip()
            if not contact:
                continue
            txn_type = (txn.get("txn_type") or "").lower()
            # Exclude non-sales AR movements: payments, refunds, adjustments.
            if txn_type in {"customer_payment", "refund", "credit_note", "creditnote"}:
                continue
            amount = float(txn.get("credit") or 0) - float(txn.get("debit") or 0)
            if amount <= 0:
                # Try the opposite sign — AR-driven path has debits, not credits.
                amount = float(txn.get("debit") or 0) - float(txn.get("credit") or 0)
            if amount <= 0:
                continue
            buckets[contact] = buckets.get(contact, 0.0) + amount
            total += amount
        return buckets, total

    def _is_sales(t: dict) -> bool:
        acct = (t.get("account_name") or "").lower()
        return any(k in acct for k in ("sales", "revenue", "income"))

    def _is_ar(t: dict) -> bool:
        acct = (t.get("account_name") or "").lower()
        if not ("accounts receivable" in acct or "sundry debtor" in acct or acct == "debtors"):
            return False
        # AR debit = invoice raised on customer. Credits are receipts/refunds.
        return float(t.get("debit") or 0) > 0 and float(t.get("credit") or 0) == 0

    buckets, total = _bucket(_is_sales)
    if not buckets:
        buckets, total = _bucket(_is_ar)

    ranked = sorted(buckets.items(), key=lambda kv: kv[1], reverse=True)[:top_n]
    return [
        {
            "customer": name,
            "revenue": amount,
            "share_pct": round(amount / total * 100, 2) if total else 0.0,
        }
        for name, amount in ranked
    ]

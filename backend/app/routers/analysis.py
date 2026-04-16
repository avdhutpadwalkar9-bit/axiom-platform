import csv
import io
import json

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from openpyxl import load_workbook
from pydantic import BaseModel

from app.middleware.auth import get_current_user
from app.models.user import User
from app.services.audited_fs_parser import parse_audited_fs_pdf
from app.services.gl_parser import customer_concentration, parse_gl_excel
from app.services.tb_analyzer import analyze_trial_balance

router = APIRouter(prefix="/api/analysis", tags=["analysis"])

# Input modes accepted by the uploader. Keep in sync with frontend's
# `InputMode` union and backend `AnalysisResult.input_mode`.
_VALID_INPUT_MODES = {"TB", "AUDITED", "GL", "PNL_ONLY", "BS_ONLY", "MIS", "SIMPLE"}


class TBEntryInput(BaseModel):
    account_code: str = ""
    account_name: str
    debit: float = 0
    credit: float = 0


class TBAnalyzeRequest(BaseModel):
    entries: list[TBEntryInput]


@router.post("/tb/json")
async def analyze_tb_json(
    data: TBAnalyzeRequest,
    current_user: User = Depends(get_current_user),
):
    """Analyze a trial balance submitted as JSON."""
    if not data.entries:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No entries provided")

    entries = [e.model_dump() for e in data.entries]
    result = analyze_trial_balance(entries)
    return result


@router.post("/tb/multi-upload")
async def analyze_multi_upload(
    files: list[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
):
    """Analyze multiple financial files (TB for different years, GL, etc.)."""
    if not files:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No files provided")

    results = []
    for file in files:
        if not file.filename:
            continue
        content = await file.read()

        entries = []
        text = ""
        if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
            text = content.decode("utf-8-sig")

        # Reuse the same parsing logic (simplified - delegates to single upload internally)
        if file.filename.endswith(".csv"):
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                account_name = (
                    row.get("account_name") or row.get("Account Name") or
                    row.get("Particulars") or row.get("Account") or ""
                )
                debit = _parse_number(row.get("debit") or row.get("Debit") or row.get("Dr") or "0")
                credit = _parse_number(row.get("credit") or row.get("Credit") or row.get("Cr") or "0")
                if account_name:
                    entries.append({"account_name": account_name.strip(), "debit": debit, "credit": credit})

        elif file.filename.endswith(".xlsx") or file.filename.endswith(".xls"):
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            if ws:
                # Smart header detection
                header_row = -1
                name_col = debit_col = credit_col = -1
                header_names = {"account", "account name", "particulars", "ledger", "name"}
                debit_names = {"debit", "net debit", "dr"}
                credit_names = {"credit", "net credit", "cr"}
                section_names = {"assets", "liabilities", "equities", "equity", "income", "revenue", "expense", "expenses"}

                for row_idx in range(1, min(11, ws.max_row + 1)):
                    cells = [str(c.value or "").strip().lower() for c in ws[row_idx]]
                    for i, cell in enumerate(cells):
                        if cell in header_names and name_col == -1: name_col = i; header_row = row_idx
                        if cell in debit_names and debit_col == -1: debit_col = i
                        if cell in credit_names and credit_col == -1: credit_col = i
                    if name_col != -1 and debit_col != -1: break

                if header_row == -1:
                    header_row = 1; name_col = 0; debit_col = 2; credit_col = 3

                current_group = ""
                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    cells = list(row)
                    raw_name = str(cells[name_col] or "").strip() if name_col < len(cells) else ""
                    if not raw_name: continue
                    clean = raw_name.lower().strip()
                    if clean in section_names: current_group = raw_name; continue
                    if clean.startswith("total for") or clean == "total": continue
                    debit = _parse_number(str(cells[debit_col] if debit_col < len(cells) else 0))
                    credit = _parse_number(str(cells[credit_col] if credit_col < len(cells) else 0))
                    if debit == 0 and credit == 0: continue
                    entries.append({"account_name": raw_name, "debit": debit, "credit": credit, "group": current_group})

        if entries:
            result = analyze_trial_balance(entries)
            results.append({
                "filename": file.filename,
                "analysis": result,
                "entry_count": len(entries),
            })

    if not results:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No valid data found in uploaded files")

    return {
        "file_count": len(results),
        "results": results,
        "combined_summary": {
            "total_files": len(results),
            "filenames": [r["filename"] for r in results],
        }
    }


@router.post("/tb/upload")
async def analyze_tb_upload(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Analyze a trial balance uploaded as CSV or JSON file."""
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No file provided")

    content = await file.read()

    entries = []

    # For text-based formats, decode content
    text = ""
    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
        text = content.decode("utf-8-sig")  # Handle BOM in Excel exports

    if file.filename.endswith(".json"):
        try:
            data = json.loads(text)
            if isinstance(data, list):
                entries = data
            elif isinstance(data, dict) and "entries" in data:
                entries = data["entries"]
            else:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid JSON format")
        except json.JSONDecodeError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid JSON file")

    elif file.filename.endswith(".csv"):
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            # Try common column name variations
            account_name = (
                row.get("account_name") or row.get("Account Name") or
                row.get("AccountName") or row.get("Particulars") or
                row.get("particulars") or row.get("Ledger") or
                row.get("ledger") or row.get("Account") or
                row.get("account") or ""
            )
            account_code = (
                row.get("account_code") or row.get("Account Code") or
                row.get("Code") or row.get("code") or ""
            )
            debit = _parse_number(
                row.get("debit") or row.get("Debit") or row.get("Dr") or row.get("dr") or "0"
            )
            credit = _parse_number(
                row.get("credit") or row.get("Credit") or row.get("Cr") or row.get("cr") or "0"
            )

            if account_name:
                entries.append({
                    "account_code": account_code,
                    "account_name": account_name,
                    "debit": debit,
                    "credit": credit,
                })
    elif file.filename.endswith(".xlsx") or file.filename.endswith(".xls"):
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            if ws is None:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Empty Excel file")

            # Smart header detection — scan first 10 rows to find the header row
            header_row = -1
            name_col = -1
            debit_col = -1
            credit_col = -1
            code_col = -1

            header_names = {"account", "account name", "particulars", "ledger", "name"}
            debit_names = {"debit", "net debit", "dr", "debit amount"}
            credit_names = {"credit", "net credit", "cr", "credit amount"}
            code_names = {"account code", "code", "accountcode"}

            for row_idx in range(1, min(11, ws.max_row + 1)):
                cells = [str(c.value or "").strip().lower() for c in ws[row_idx]]
                for i, cell in enumerate(cells):
                    if cell in header_names and name_col == -1:
                        name_col = i
                        header_row = row_idx
                    if cell in debit_names and debit_col == -1:
                        debit_col = i
                    if cell in credit_names and credit_col == -1:
                        credit_col = i
                    if cell in code_names and code_col == -1:
                        code_col = i
                if name_col != -1 and debit_col != -1:
                    break

            if header_row == -1:
                # Fallback: assume col 0=name, 2=debit, 3=credit (common Zoho/Tally format)
                header_row = 1
                name_col = 0
                debit_col = 2
                credit_col = 3

            # Track current group/section from Zoho-style section headers
            current_group = ""
            skip_prefixes = {"total for", "total"}
            section_names = {"assets", "liabilities", "equities", "equity", "income", "revenue", "expense", "expenses"}

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                cells = list(row)
                if name_col >= len(cells):
                    continue

                raw_name = str(cells[name_col] or "").strip()
                if not raw_name:
                    continue

                # Check if this is a section header (Assets, Liabilities, Income, Expense)
                clean_name = raw_name.lower().strip()
                if clean_name in section_names:
                    current_group = raw_name
                    continue

                # Skip "Total for..." rows
                if clean_name.startswith("total for") or clean_name == "total":
                    continue

                # Skip the header row itself if it appears in data
                if clean_name in header_names:
                    continue

                debit = _parse_number(str(cells[debit_col] if debit_col >= 0 and debit_col < len(cells) else 0))
                credit = _parse_number(str(cells[credit_col] if credit_col >= 0 and credit_col < len(cells) else 0))
                account_code = str(cells[code_col] if code_col >= 0 and code_col < len(cells) else "").strip()

                # Skip rows with no debit AND no credit (empty rows or zero-balance accounts)
                if debit == 0 and credit == 0:
                    continue

                # Prefix the account name with the current group for better classification
                # e.g., "Sales" under "Income" group becomes more easily classifiable
                group_hint = ""
                if current_group:
                    group_hint = current_group.strip()

                entries.append({
                    "account_code": account_code,
                    "account_name": raw_name.strip(),
                    "debit": debit,
                    "credit": credit,
                    "group": group_hint,
                })

        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Error reading Excel file: {str(e)}")

    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file format. Please upload a .csv, .json, or .xlsx file."
        )

    if not entries:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No valid entries found in file")

    result = analyze_trial_balance(entries)
    return result


class QuickAnalysisRequest(BaseModel):
    company_name: str = ""
    industry: str = ""
    turnover_range: str = ""
    entries: list[TBEntryInput]


@router.post("/tb/quick")
async def quick_analysis(data: QuickAnalysisRequest):
    """
    Quick analysis WITHOUT authentication — the landing page hook.
    Returns ONE painful insight + blurred summary to force signup.
    """
    if not data.entries or len(data.entries) < 3:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Need at least 3 entries")

    entries = [e.model_dump() for e in data.entries]
    full_result = analyze_trial_balance(entries)

    # Extract the ONE most painful insight
    fs = full_result["financial_statements"]
    ratios = full_result["ratios"]
    insights = full_result["insights"]

    # Find the most critical insight
    painful_insight = None
    for ins in insights:
        if ins["severity"] == "critical":
            painful_insight = ins
            break
    if not painful_insight and insights:
        painful_insight = insights[0]

    # Build the teaser response — only partial data
    industry_benchmarks = {
        "Manufacturing": {"gross_margin": 35, "current_ratio": 1.5, "label": "Manufacturing"},
        "SaaS": {"gross_margin": 70, "current_ratio": 2.0, "label": "SaaS/Tech"},
        "Services": {"gross_margin": 45, "current_ratio": 1.3, "label": "Services"},
        "Trading": {"gross_margin": 20, "current_ratio": 1.2, "label": "Trading"},
        "E-commerce": {"gross_margin": 30, "current_ratio": 1.4, "label": "E-commerce"},
    }
    benchmark = industry_benchmarks.get(data.industry, {"gross_margin": 40, "current_ratio": 1.5, "label": "your industry"})

    # Generate the ONE painful line
    pain_lines = []
    if ratios["current_ratio"] > 0 and ratios["current_ratio"] < benchmark["current_ratio"]:
        wc_leak = abs(ratios["working_capital"])
        pain_lines.append(f"We found a potential ₹{wc_leak/100000:.1f}L working capital leak — your current ratio ({ratios['current_ratio']}x) is below the {benchmark['label']} benchmark of {benchmark['current_ratio']}x.")
    if ratios["gross_margin"] > 0 and ratios["gross_margin"] < benchmark["gross_margin"]:
        gap = benchmark["gross_margin"] - ratios["gross_margin"]
        pain_lines.append(f"Your Gross Margin ({ratios['gross_margin']}%) is {gap:.1f}% below the {benchmark['label']} industry average.")
    if ratios["net_margin"] < 0:
        pain_lines.append(f"Your business is operating at a {abs(ratios['net_margin'])}% loss. Immediate cost optimization needed.")
    if fs["net_income"] > 0 and ratios["net_margin"] < 10:
        pain_lines.append(f"Net margin of {ratios['net_margin']}% is thin — one bad quarter could push you into losses.")

    if not pain_lines:
        pain_lines.append(f"Revenue of ₹{fs['total_revenue']/100000:.1f}L with {ratios['net_margin']}% margin — there are optimization opportunities we've identified.")

    return {
        "teaser": {
            "headline": pain_lines[0],
            "supporting": pain_lines[1] if len(pain_lines) > 1 else None,
            "company_name": data.company_name or "Your Company",
        },
        "summary_preview": {
            "total_revenue": fs["total_revenue"],
            "total_expenses": fs["total_expenses"],
            "net_income": fs["net_income"],
            "account_count": len(entries),
        },
        "blurred": {
            "insights_count": len(insights),
            "ind_as_issues": len(full_result["ind_as_observations"]),
            "ai_questions": len(full_result["ai_questions"]),
            "ratios_available": True,
            "balance_sheet_available": True,
        },
        "cta": "Sign up to unlock the full Ind AS review, 5 AI questions, and actionable SOPs →",
    }


def _parse_number(value: str) -> float:
    """Parse a number from string, handling commas and parentheses."""
    if not value:
        return 0.0
    value = str(value).strip()
    # Handle parentheses as negative
    if value.startswith("(") and value.endswith(")"):
        value = "-" + value[1:-1]
    # Remove commas and currency symbols
    value = value.replace(",", "").replace("₹", "").replace("$", "").replace(" ", "")
    try:
        return float(value)
    except ValueError:
        return 0.0


# =========================================================================
# Unified upload dispatcher — picks the right parser based on `input_mode`
# and feeds the resulting synthetic TB into `analyze_trial_balance` so the
# rest of the pipeline (ratios_meta, Ind AS checks, insights) is shared.
#
# This is deliberately additive; the legacy /tb/upload and /tb/json routes
# above stay untouched so old callers keep working.
# =========================================================================


def _parse_tb_file(file_name: str, content: bytes) -> list[dict]:
    """Parse a TB-shaped file (CSV / JSON / Excel) into raw entries.

    Extracted so both the legacy /tb/upload route and the new dispatcher
    can share the code. We re-run the light-weight header detection used
    by /tb/upload — for brevity we just redirect heavy-lifting callers
    to the existing logic when needed.
    """
    entries: list[dict] = []
    name_l = file_name.lower()

    if name_l.endswith(".json"):
        text = content.decode("utf-8-sig")
        try:
            data = json.loads(text)
        except json.JSONDecodeError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid JSON file: {exc}",
            )
        if isinstance(data, list):
            entries = data
        elif isinstance(data, dict) and "entries" in data:
            entries = data["entries"]
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Invalid JSON format — expected list or {entries: [...]}.",
            )
        return entries

    if name_l.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            account_name = (
                row.get("account_name") or row.get("Account Name") or
                row.get("AccountName") or row.get("Particulars") or
                row.get("Account") or row.get("Ledger") or ""
            )
            debit = _parse_number(
                row.get("debit") or row.get("Debit") or row.get("Dr") or "0"
            )
            credit = _parse_number(
                row.get("credit") or row.get("Credit") or row.get("Cr") or "0"
            )
            if account_name:
                entries.append({
                    "account_name": account_name.strip(),
                    "debit": debit,
                    "credit": credit,
                })
        return entries

    if name_l.endswith(".xlsx") or name_l.endswith(".xls"):
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Empty Excel file",
            )

        header_row = -1
        name_col = debit_col = credit_col = code_col = -1
        header_names = {"account", "account name", "particulars", "ledger", "name"}
        debit_names = {"debit", "net debit", "dr", "debit amount"}
        credit_names = {"credit", "net credit", "cr", "credit amount"}
        code_names = {"account code", "code", "accountcode"}

        for row_idx in range(1, min(11, ws.max_row + 1)):
            cells = [str(c.value or "").strip().lower() for c in ws[row_idx]]
            for i, cell in enumerate(cells):
                if cell in header_names and name_col == -1:
                    name_col = i
                    header_row = row_idx
                if cell in debit_names and debit_col == -1:
                    debit_col = i
                if cell in credit_names and credit_col == -1:
                    credit_col = i
                if cell in code_names and code_col == -1:
                    code_col = i
            if name_col != -1 and debit_col != -1:
                break

        if header_row == -1:
            header_row = 1
            name_col = 0
            debit_col = 2
            credit_col = 3

        current_group = ""
        section_names = {
            "assets", "liabilities", "equities", "equity",
            "income", "revenue", "expense", "expenses",
        }
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            cells = list(row)
            if name_col >= len(cells):
                continue
            raw_name = str(cells[name_col] or "").strip()
            if not raw_name:
                continue
            clean_name = raw_name.lower().strip()
            if clean_name in section_names:
                current_group = raw_name
                continue
            if clean_name.startswith("total for") or clean_name == "total":
                continue
            if clean_name in header_names:
                continue
            debit = _parse_number(
                str(cells[debit_col] if 0 <= debit_col < len(cells) else 0)
            )
            credit = _parse_number(
                str(cells[credit_col] if 0 <= credit_col < len(cells) else 0)
            )
            code = str(cells[code_col] if 0 <= code_col < len(cells) else "").strip()
            if debit == 0 and credit == 0:
                continue
            entries.append({
                "account_code": code,
                "account_name": raw_name,
                "debit": debit,
                "credit": credit,
                "group": current_group,
            })
        return entries

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=(
            f"Unsupported file format for TB mode: {file_name}. "
            "Please upload .csv, .json, .xlsx or .xls."
        ),
    )


@router.post("/upload")
async def analyze_upload(
    file: UploadFile = File(...),
    input_mode: str = Form("TB"),
    current_user: User = Depends(get_current_user),
):
    """Unified upload endpoint.

    Accepts an `input_mode` form field telling the server what to expect:
      - "TB" (default): trial balance (CSV / JSON / XLSX)
      - "AUDITED": Indian Schedule III audited financial statements (PDF)
      - "GL": General Ledger export (Zoho Books / Tally / QuickBooks XLSX)
      - "PNL_ONLY", "BS_ONLY", "MIS", "SIMPLE": routed through the TB path
        today (the analyzer copes with partial data thanks to
        ratios_meta); dedicated parsers will land in later waves.

    Returns the same shape as `/tb/upload`, with an extra `upload_meta`
    block when a richer parser is used (GL transactions, unit multipliers,
    parser warnings, etc.)."""
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No file provided",
        )
    mode = (input_mode or "TB").upper().strip()
    if mode not in _VALID_INPUT_MODES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Unknown input_mode '{input_mode}'. "
                f"Expected one of: {sorted(_VALID_INPUT_MODES)}"
            ),
        )

    content = await file.read()
    upload_meta: dict = {"file_name": file.filename, "declared_mode": mode}

    # -------- AUDITED: PDF (or Excel export) --------
    if mode == "AUDITED":
        if not file.filename.lower().endswith(".pdf"):
            # Let Excel exports of audited statements fall through to the
            # TB parser — still tagged as AUDITED in the final result.
            entries = _parse_tb_file(file.filename, content)
            upload_meta["parser"] = "tb_fallback_for_audited_excel"
        else:
            parsed = parse_audited_fs_pdf(content)
            entries = parsed.tb_entries
            upload_meta.update({
                "parser": "audited_fs_pdf",
                "pages_parsed": parsed.pages_parsed,
                "unit_multiplier": parsed.unit_multiplier,
                "raw_text_length": parsed.raw_text_length,
                "parser_warnings": parsed.warnings,
            })

    # -------- GL: transaction-level ledger --------
    elif mode == "GL":
        if not (file.filename.lower().endswith((".xlsx", ".xls"))):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "GL mode expects an Excel export (.xlsx / .xls) from "
                    "Zoho Books, Tally or QuickBooks. Please re-export the "
                    "General Ledger and retry."
                ),
            )
        parsed = parse_gl_excel(content)
        entries = parsed.tb_entries
        upload_meta.update({
            "parser": "gl_parser",
            "source_format": parsed.source_format,
            "transaction_count": parsed.total_rows,
            "parser_warnings": parsed.warnings,
            "customer_concentration": customer_concentration(parsed.transactions),
        })

    # -------- TB / PNL_ONLY / BS_ONLY / MIS / SIMPLE --------
    else:
        entries = _parse_tb_file(file.filename, content)
        upload_meta["parser"] = "tb_generic"

    if not entries:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                upload_meta.get("parser_warnings", [
                    "No valid entries found in file"
                ])[0]
                if upload_meta.get("parser_warnings")
                else "No valid entries found in file"
            ),
        )

    result = analyze_trial_balance(entries)
    # Overwrite the analyzer's default "TB" tag with the user-declared mode
    # so the frontend knows what it's looking at.
    result["input_mode"] = mode
    result["upload_meta"] = upload_meta
    return result
